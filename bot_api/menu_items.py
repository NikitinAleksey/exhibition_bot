import os
from copy import copy
from datetime import datetime, timedelta
from typing import Dict, Optional, List, Callable, Literal, Union
from urllib.parse import urlparse

import openpyxl
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton

from avito_api.avito import get_items_stats, get_autoload_last_completed_report
from database.db import (get_all_info_about_company, get_in_charge_admin_ids, get_admin_names,
                         get_companies_titles_list, get_admin_ids, get_links_to_chats, insert_record_to_common_table,
                         delete_company, delete_record_from_customers_with_condition, insert_admin,
                         delete_admin_with_name)
from bot_api.server_requests import Request
from bot_api.keyboards import KeyboardManager
from log_settings.logger_init import logger


class BaseMethodsAndData:
    yes_no_buttons: List = ['Да', 'Нет']
    yes_no_callbacks: List = ['yes', 'no']
    back_button: str = '⬅️ Назад'
    back_customers_menu_callback: str = 'customers_menu'

    @classmethod
    def create_inline_keyboard(cls, buttons: List[str], callbacks: List[str]) -> InlineKeyboardMarkup:
        """
        Возвращает ин лайн клавиатуру указанного меню.
        Кнопки и коллбэки лучше брать отсюда или из класса-наследника, например:
        keyboard = Customer.create_inline_keyboard(
                                         buttons=Customer.autoload_menu_buttons,
                                         callbacks=Customer.autoload_menu_callbacks
                                                 )
        """
        return KeyboardManager(buttons=buttons, callbacks=callbacks).create_inline_keyboard()

    @classmethod
    def create_inline_buttons(cls,
                              buttons: List[str],
                              callbacks: Optional[List[str]] = None,
                              urls: Optional[List[str]] = None) -> List[List[InlineKeyboardButton]]:
        """Создаст либо обычные кнопки, либо кнопки со ссылками в зависимости от того, что передать в метод."""

        if callbacks:
            callback_buttons = [
                [InlineKeyboardButton(text=button, callback_data=callback)]
                for button, callback in zip(buttons, callbacks)
            ]
            return callback_buttons

        elif urls:
            url_buttons = [
                [InlineKeyboardButton(text=button, url=url)]
                for button, url in zip(buttons, urls)
            ]
            return url_buttons


class MainMenu:
    pass


class EmployeeMenu(BaseMethodsAndData):
    employee_menu_header: str = ('Меню сотрудников.\n'
                                 'Выберите пункт меню:')
    employee_menu_buttons: List = ['Меню клиентов', 'Админка', 'Скрыть меню']
    employee_menu_callbacks: List = ['customers_menu', 'admin_menu', 'close_menu']
    employee_menu_restriction: str = 'Доступ запрещен.'

    @classmethod
    def is_employee(cls, telegram_id: int) -> bool:
        return telegram_id in get_admin_ids()


class CustomersMenu(BaseMethodsAndData):
    customer_menu_header: str = ('Меню клиентов.\n'
                                 'Выберите пункт меню:')
    customer_menu_buttons: List = ['Все клиенты', 'Добавить клиента', 'Удалить клиента', 'Изменить клиента',
                                   '⬅️ Назад']
    customer_menu_callbacks: List = ['all_customers', 'add_customer', 'delete_customer', 'edit_customer',
                                     'employee_menu']


class AllCustomers(BaseMethodsAndData):
    all_customers_header: str = 'Выберите клиента:'
    all_customers_buttons: List = []
    all_customers_callbacks: List = []

    @classmethod
    def update_buttons(cls):
        cls.all_customers_buttons = get_companies_titles_list()
        cls.all_customers_buttons.append(cls.back_button)
        return cls.all_customers_buttons

    @classmethod
    def update_callbacks(cls):
        cls.all_customers_callbacks = cls.all_customers_buttons[:-1]
        cls.all_customers_callbacks.append(cls.back_customers_menu_callback)
        return cls.all_customers_callbacks


class AdminMenu(BaseMethodsAndData):
    admin_menu_header: str = 'Меню админа.\nВыберите действие:'
    is_correct: str = 'Введенные данные верны?'
    adding_new_admin_name = 'Введите имя:'
    adding_new_admin_id = 'Введите телеграмм ID:'
    admin_name_to_delete = 'Выберите администратора на удаление:'
    back_to_admin_menu_callback = 'admin_menu'
    back_to_employee_menu_callback: str = 'employee_menu'
    new_admin_is_added: str = 'Новый админ {admin_name} добавлен'
    new_admin_is_not_added: str = ('Произошла ошибка при добавлении {admin_name}:'
                                   '{exc}')

    admin_menu_buttons: List[str] = ['Добавить сотрудника', 'Удалить сотрудника', BaseMethodsAndData.back_button]
    admin_menu_callbacks: List[str] = ['add_employee', 'delete_employee', back_to_employee_menu_callback]
    admin_menu_restriction: str = 'Доступ запрещен.'
    admin_is_deleted: str = 'Админ {admin_name} успешно удален'
    admin_is_not_deleted: str = 'Админа {admin_name} невозможно удалить'
    not_valid_tg_id: str = ('Телеграм ID должен быть числом.\n'
                            'Введите Телеграм ID:')
    admin_in_charge_ids: Callable = get_in_charge_admin_ids
    admin_names_buttons: List[str] = []
    admin_names_callbacks: List[str] = []

    def __init__(self, admin_name: str) -> None:
        self.admin_name = admin_name
        self.admin_id = None

    def set_admin_id(self, tg_id: str) -> str:
        self.admin_id = tg_id
        return self.admin_id

    def add_new_admin_to_db(self) -> str:
        try:
            insert_admin(admin_id=self.admin_id, admin_name=self.admin_name)
            return self.__class__.new_admin_is_added.format(
                admin_name=self.admin_name
            )
        except Exception as exc:
            logger.warning(msg=f'Ошибка при добавлении нового админа: {exc}')
            return self.__class__.new_admin_is_not_added.format(
                admin_name=self.admin_name,
                exc=exc
            )

    @classmethod
    def update_buttons(cls):
        cls.admin_names_buttons = get_admin_names()
        cls.admin_names_buttons.append(cls.back_button)
        return cls.admin_names_buttons

    @classmethod
    def update_callbacks(cls):
        cls.admin_names_callbacks = cls.admin_names_buttons[:-1]
        cls.admin_names_callbacks.append(cls.back_to_admin_menu_callback)
        return cls.admin_names_callbacks

    @classmethod
    def is_admin(cls, telegram_id: int) -> bool:
        return telegram_id in get_in_charge_admin_ids()

    @classmethod
    def delete_admin_if_not_in_charge(cls, admin_name: str) -> str:
        if delete_admin_with_name(admin_name=admin_name):
            return cls.admin_is_deleted.format(admin_name=admin_name)
        return cls.admin_is_not_deleted.format(admin_name=admin_name)

    @classmethod
    def validate_tg_id(cls, tg_id: str) -> bool:
        return tg_id.isdigit()


class AddCustomer(BaseMethodsAndData):
    common_header: str = 'Введите {item}: '
    title_item: str = 'название'
    avito_id_item: str = 'Avito ID'
    client_id_item: str = 'Client ID'
    client_secret_item: str = 'Client Secret'
    chat_with_client_item: str = 'Чат с клиентом'
    chat_about_client_item: str = 'Чат по клиенту'
    google_doc_link: str = 'Ссылка на Google doc'

    not_valid_avito_id: str = 'Avito ID должен быть числом. Попробуйте еще раз.'
    not_valid_link: str = 'Это должна быть ссылка. Попробуйте еще раз.'

    successfully_added_customer: str = 'Клиент {customer} успешно внесен в базу данных'
    unsuccessfully_added_customer: str = 'Возникла ошибка при внесении клиента: {error}'

    def __init__(self, title: str) -> None:
        self.title = title
        self.avito_id = None
        self.client_id = None
        self.client_secret = None
        self.chat_with_client_link = None
        self.chat_about_client_link = None
        self.google_doc_link = None

    def set_avito_id(self, avito_id: str) -> Optional[str]:
        if not avito_id.isdigit():
            return
        self.avito_id = avito_id
        return self.avito_id

    def set_client_id(self, client_id: str) -> None:
        self.client_id = client_id

    def set_client_secret(self, client_secret: str) -> None:
        self.client_secret = client_secret

    def set_chat_with_client_link(self, chat_with_client_link: str) -> bool:
        try:
            result = urlparse(chat_with_client_link)
            self.chat_with_client_link = chat_with_client_link
            return all([result.scheme, result.netloc])
        except ValueError as exc:
            logger.info(msg=f'Неверный формат ссылки: {exc}')
            return False

    def set_chat_about_client_link(self, chat_about_client_link: str) -> bool:
        try:
            result = urlparse(chat_about_client_link)
            self.chat_about_client_link = chat_about_client_link
            return all([result.scheme, result.netloc])
        except ValueError as exc:
            logger.info(msg=f'Неверный формат ссылки: {exc}')
            return False

    def set_google_doc_link(self, google_doc_link: str) -> bool:
        try:
            result = urlparse(google_doc_link)
            self.google_doc_link = google_doc_link
            return all([result.scheme, result.netloc])
        except ValueError as exc:
            logger.info(msg=f'Неверный формат ссылки: {exc}')

            return False

    def check_data_items(self) -> str:
        is_correct: str = (f'Введенные данные верны?\n'
                           f'Название: {self.title}\n'
                           f'Avito ID: {self.avito_id}\n'
                           f'Client ID: {self.client_id}\n'
                           f'Client Secret: {self.client_secret}\n'
                           f'Чат с клиентом: {self.chat_with_client_link}\n'
                           f'Чат по клиенту: {self.chat_about_client_link}\n'
                           f'Ссылка на Google doc: {self.google_doc_link}')
        return is_correct

    def add_customer_to_db(self) -> str:
        try:
            insert_record_to_common_table(
                company_name=self.title,
                avito_id=self.avito_id,
                client_id=self.client_id,
                client_secret=self.client_secret,
                chat_with_client=self.chat_with_client_link,
                chat_about_client=self.chat_about_client_link,
                google_doc_link=self.google_doc_link
            )
            return self.successfully_added_customer.format(customer=self.title)
        except Exception as exc:
            logger.warning(msg=f'Ошибка при добавлении нового клиента: {exc}')
            return self.unsuccessfully_added_customer.format(error=exc)


class DeleteCustomer(BaseMethodsAndData):
    delete_customer_header: str = 'Выберите клиента для удаления: '
    delete_customer_buttons: List[str] = []
    delete_customer_callbacks: List[str] = []

    customer_is_deleted: str = 'Клиент {title} успешно удален'
    customer_is_not_deleted: str = 'При удалении клиента произошла ошибка {error}'

    @classmethod
    def update_buttons(cls) -> List[str]:
        cls.delete_customer_buttons = get_companies_titles_list()
        cls.delete_customer_buttons.append(cls.back_button)
        return cls.delete_customer_buttons

    @classmethod
    def update_callbacks(cls) -> List[str]:
        cls.delete_customer_callbacks = cls.delete_customer_buttons[:-1]
        cls.delete_customer_callbacks.append(cls.back_customers_menu_callback)
        return cls.delete_customer_callbacks

    @classmethod
    def delete_customer(cls, title: str):
        try:
            delete_company(company_name=title)
            return cls.customer_is_deleted.format(title=title)
        except Exception as exc:
            logger.warning(msg=f'Ошибка при удалении клиента: {exc}')
            return cls.customer_is_not_deleted.format(error=exc)


class EditCustomer(BaseMethodsAndData):
    edit_customer_header: str = 'Выберите клиента для изменения: '
    edit_attribute_header: str = 'Введите новое значение "{attribute}":'
    back_edit_customer_callback: str = 'edit_customer'
    successfully_edited_customer: str = 'Клиент {customer} успешно изменен\n'
    unsuccessfully_edited_customer: str = 'Возникла ошибка при изменении клиента: {error}\n'

    edit_customer_buttons: List[str] = []
    edit_customer_callbacks: List[str] = []

    edit_customer_fields_buttons: List[str] = [
        'Название', 'Avito ID', 'Client ID', 'Client Secret', 'Чат с клиентом', 'Чат по клиенту',
        'Ссылка на Google doc',
        BaseMethodsAndData.back_button
    ]
    edit_customer_fields_callbacks: List[str] = [
        'title', 'avito_id', 'client_id', 'client_secret', 'chat_with_client_link', 'chat_about_client_link',
        'google_doc_link', BaseMethodsAndData.back_customers_menu_callback
    ]

    def __init__(self, title: str) -> None:
        self.title = title
        self.avito_id = None
        self.client_id = None
        self.client_secret = None
        self.chat_with_client_link = None
        self.chat_about_client_link = None
        self.google_doc_link = None
        self.data: Dict = self.get_all_customer_info_from_bd()

        for key, value in self.data.items():
            setattr(self, key, value)

    def get_all_customer_info_from_bd(self):
        data = get_all_info_about_company(title=self.title)
        return data

    def represent_current_data(self, already_changed: bool = False, is_success: Union[Optional[str]] = None) -> str:
        if not already_changed:
            represent_header: str = 'Что вы хотели бы изменить?\n'
        else:
            represent_header: str = 'Новые данные по клиенту:\n'
            if is_success:
                represent_header = is_success + represent_header
        attributes_list = [
            self.title, self.avito_id, self.client_id, self.client_secret, self.chat_with_client_link,
            self.chat_about_client_link, self.google_doc_link
        ]
        for field, data in zip(self.__class__.edit_customer_fields_buttons, attributes_list):
            temp_string = field + ': ' + str(data) + '\n'
            represent_header += temp_string
        return represent_header

    @classmethod
    def update_buttons(cls) -> List[str]:
        cls.edit_customer_buttons = get_companies_titles_list()
        cls.edit_customer_buttons.append(cls.back_button)
        return cls.edit_customer_buttons

    @classmethod
    def update_callbacks(cls) -> List[str]:
        cls.edit_customer_callbacks = cls.edit_customer_buttons[:-1]
        cls.edit_customer_callbacks.append(cls.back_customers_menu_callback)
        return cls.edit_customer_callbacks

    def set_chosen_attribute(self, attribute: str, value: Union[str, int]):
        ask_to_input_again = '\nПопробуйте еще раз:'
        if attribute == 'title':
            if value in get_companies_titles_list():
                return 'Такое название уже есть.' + ask_to_input_again
        elif attribute == 'avito_id':
            if not value.isdigit():
                return 'Avito ID должен быть числом.' + ask_to_input_again
        elif attribute in ['chat_with_client_link', 'chat_about_client_link', 'google_doc_link']:
            result = urlparse(value)
            if not all([result.scheme, result.netloc]):
                return 'Это должно быть ссылкой.' + ask_to_input_again
        setattr(self, attribute, value)

    @staticmethod
    def callback_data_to_field_name(attribute: str) -> str:
        data = {
            'title': 'Название',
            'avito_id': 'Avito ID',
            'client_id': 'Client ID',
            'client_secret': 'Client Secret',
            'chat_with_client_link': 'Чат с клиентом',
            'chat_about_client_link': 'Чат по клиенту',
            'google_doc_link': 'Ссылка на Google doc'
        }
        return data.get(attribute)

    def change_customer_in_db(self) -> str:
        try:
            if delete_record_from_customers_with_condition(
                    title=self.title,
                    avito_id=self.avito_id,
                    client_id=self.client_id
            ):
                insert_record_to_common_table(
                    company_name=self.title,
                    avito_id=self.avito_id,
                    client_id=self.client_id,
                    client_secret=self.client_secret,
                    chat_with_client=self.chat_with_client_link,
                    chat_about_client=self.chat_about_client_link,
                    google_doc_link=self.google_doc_link
                )
                return self.successfully_edited_customer.format(customer=self.title)
        except Exception as exc:
            logger.warning(msg=f'Ошибка при изменении клиента: {exc}')
            return self.unsuccessfully_edited_customer.format(error=exc)


class Customer(BaseMethodsAndData):
    customer_menu_button: str = '⬅️ Назад'
    customer_menu_callback: str = 'all_customers'
    customer_menu_buttons: List = ['Автозагрузка', 'Статистика']
    customer_menu_callbacks: List = ['autoload', 'statistic']
    customer_menu_returning: str = 'customer_menu'

    autoload_menu_buttons: List = ['Получить ссылку на фид', 'Удалить фид',
                                   'Обновить фид', 'Отчет по автозагрузке', customer_menu_button]
    autoload_menu_callbacks: List = ['get_link', 'delete_link',
                                     'update_feed', 'get_autoload_report', customer_menu_returning]
    autoload_menu_get_link: str = ('Ссылка на фид клиента {title}:\n'
                                   '{url}')
    autoload_menu_delete_link: str = 'Фид клиента {title} успешно удален'
    autoload_menu_back_button: str = 'autoload'
    autoload_menu_no_file_in_google_doc: str = 'Файла {file_name} нет в Google doc.'

    statistic_header: str = 'Выберите, по какому периоду времени сортировать статистику:'
    statistic_menu_buttons: List = ['День', 'Неделя', 'Месяц', customer_menu_button]
    statistic_menu_callbacks: List = ['day', 'week', 'month', customer_menu_returning]
    statistic_dates_format: str = '%Y-%m-%d'
    statistic_calendar_date_from: datetime = datetime(2020, 1, 1)
    statistic_calendar_date_to: datetime = datetime(2040, 12, 31)
    statistic_date_from: str = 'Выберите дату начала статистики:'
    statistic_date_to: str = 'Выберите дату окончания статистики:'
    statistic_dates_choice: str = ('Выбранные даты верны?\n'
                                   'Начало статистики: {date_from}\n'
                                   'Окончание статистики: {date_to}')
    statistic_back_callback: str = 'statistic'
    statistic_empty_data: str = 'Статистика пуста. Невозможно сформировать файл.'
    statistic_access_problem: str = ('Проблема с доступом. Проверьте корректность Client ID и Client Secret.\n'
                                     'Если все указано верно, значит статистика пустая (по всем параметрам нули).\n'
                                     'Выберите другие даты для статистики.')

    def __init__(self, title: str) -> None:
        self.title = title
        self.data: Dict = self.get_data_about_customer()
        self.avito_id = None
        self.client_id = None
        self.client_secret = None
        self.chat_with_client = None
        self.chat_about_client = None

        for key, value in self.data.items():
            setattr(self, key, value)

    def header_customer(self, main: bool = False, autoload: bool = False, statistics: bool = False) -> str:
        """
        Вернет строку для ответа пользователю, которая будет над кнопками.
        Нужно указать нужное при вызове функции.
        """
        common_customer_info = (f"Клиент: {self.title}\n"
                                f"Avito ID: {self.avito_id}\n")
        if main:
            option = 'клиенту'
        elif autoload:
            option = 'автозагрузке'
        elif statistics:
            option = 'статистике'
        else:
            option = ''
        return common_customer_info + f'Выберите действие по {option}:\n'

    def get_data_about_customer(self) -> Dict:
        """
        Вернет словарь со всеми значащими элементами из таблицы базы данных:
        - avito_id,
        - title,
        - client_id,
        - client_secret,
        - chat_with_client_link,
        - chat_about_client_link

        :return: Dict.
        """
        return get_all_info_about_company(title=self.title)

    # Раздел блока автозагрузки
    def get_autoload_link(self) -> Optional[str]:
        """Вернет либо ссылку на фид, либо None, если ссылки не существует или возникла ошибка."""
        return Request.get_url_to_feed(company_name=self.title)

    def delete_autoload_link(self) -> Optional[str]:
        """Вернет строку с информацией об успешном удалении, либо об ошибках при запросе."""
        return Request.delete_file(company_name=self.title)

    # Раздел блока статистики
    def get_statistic(self, date_from: str, date_to: str, period: Literal['week', 'month', 'year']):
        try:
            return get_items_stats(
                company_name=self.title,
                date_from=date_from,
                date_to=date_to,
                period=period
            )
        except KeyError as exc:
            logger.warning(msg=f'Ошибка при получении статистики: {exc}')
            return

    def create_exel_file_from_statistic(self, data: List[Dict]):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet['A1'] = 'Номер объявления'
        sheet['B1'] = 'Дата'
        sheet['C1'] = 'Контакты'
        sheet['D1'] = 'Избранные'
        sheet['E1'] = 'Просмотры'

        # Заполнение данными
        row = 2
        try:
            for item in data:
                item_id = item['itemId']
                stats = item['stats']

                start_row = row
                end_row = row + len(stats) - 1
                sheet.cell(row=start_row, column=1, value=item_id)

                if stats:
                    sheet.merge_cells(f'A{start_row}:A{end_row}')

                for entry in stats:
                    date = datetime.strptime(entry['date'], '%Y-%m-%d').date().isoformat()
                    uniq_contacts = entry['uniqContacts']
                    uniq_favorites = entry['uniqFavorites']
                    uniq_views = entry['uniqViews']

                    sheet.cell(row=row, column=2, value=date)
                    sheet.cell(row=row, column=3, value=uniq_contacts)
                    sheet.cell(row=row, column=4, value=uniq_favorites)
                    sheet.cell(row=row, column=5, value=uniq_views)

                    row += 1

            filename = f'{self.title}.xlsx'
            workbook.save(filename)

            return filename

        except TypeError as exc:
            logger.warning(msg=f'Ошибка при создании эксель файла: {exc}')
            return exc

    @classmethod
    def delete_temp_exel_file(cls, filename: str) -> None:
        os.remove(filename)

    # Раздел для ссылок на чаты мессенджера
    def get_links_to_customer_chats(self):
        return get_links_to_chats(title=self.title)

    def create_inline_keyboard_with_urls(self) -> InlineKeyboardMarkup:
        buttons = copy(self.__class__.customer_menu_buttons)
        callbacks = copy(self.__class__.customer_menu_callbacks)
        customer_chats_buttons = ['Чат с клиентом', 'Чат по клиенту', 'Ссылка на Google doc']
        customer_chats_urls = self.get_links_to_customer_chats()
        back_button = [self.__class__.customer_menu_button]
        back_callback = [self.__class__.customer_menu_callback]

        data_buttons = self.__class__.create_inline_buttons(buttons=buttons, callbacks=callbacks)
        url_buttons = self.__class__.create_inline_buttons(buttons=customer_chats_buttons, urls=customer_chats_urls)
        back_buttons = self.__class__.create_inline_buttons(buttons=back_button, callbacks=back_callback)

        all_buttons = data_buttons + url_buttons + back_buttons
        return InlineKeyboardMarkup(inline_keyboard=all_buttons)

    def get_autoload_report(self) -> str:
        report = get_autoload_last_completed_report(company_name=self.title)
        if report:
            try:
                string_representation_report = self.represent_dict_report_to_string(data=report)
                return string_representation_report
            except Exception as exc:
                return f'Ошибка: {exc}'
        else:
            return f'Невозможно сформировать статистику по клиенту {self.title}'

    @staticmethod
    def represent_dict_report_to_string(data: Dict) -> str:
        status = data.get('status')
        events = '\n'.join([event.get('description') for event in data.get('events')])
        started_at = Customer.time_formatter(data.get('started_at'))
        finished_at = Customer.time_formatter(data.get('finished_at'))
        total_items = data.get('section_stats').get('count')
        common_stats = (f'Статус: {status}\n'
                        f'Ошибки/предупреждения: {events}\n'
                        f'Начало загрузки: {started_at}\n'
                        f'Окончание загрузки: {finished_at}\n'
                        f'Всего объявлений: {total_items}\n')
        try:
            extracted_items_stats = Customer.extract_items_report(data=data.get('section_stats').get('sections'))
            result = common_stats + extracted_items_stats
        except Exception as exc:
            result = f'Возникла ошибка при распаковке отчета: {exc}'

        return result

    @staticmethod
    def extract_items_report(data: List[Dict]) -> str:
        pattern = '\n'.join([f'{section.get("title")}: {section.get("count")}' for section in data])

        return pattern

    @staticmethod
    def time_formatter(date: str) -> str:
        try:
            timestamp = datetime.strptime(date, "%Y-%m-%dT%H:%M:%SZ")
            msk_time_zone = timestamp + timedelta(hours=3)
            formatted_started_at = msk_time_zone.strftime("%Y-%m-%d %H:%M")
        except Exception as exc:
            logger.info(msg=f'Ошибка при форматировании латы и времени: {exc}')
            formatted_started_at = '-'
        return formatted_started_at


if __name__ == '__main__':
    pass
